# -*- coding: utf-8 -*-
"""
    th-e-yield.evaluation
    ~~~~~~~~~~~~~~~~~


"""
import os
import logging

import numpy as np
import pandas as pd
import datetime as dt

from matplotlib import pyplot as plt

from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

logger = logging.getLogger(__name__)

AC_P = 'Power [W]'
AC_E = 'Energy yield [kWh]'
AC_Y = 'Specific yield [kWh/kWp]'
DC_P = 'Power (DC) [W]'
DC_E = 'Energy yield (DC) [kWh]'


class Evaluation:

    def _evaluate_yield(self, results, weather):
        hours = pd.Series(weather.index, index=weather.index)
        hours = round((hours - hours.shift(1)).fillna(method='bfill').dt.total_seconds() / 3600)

        results_total_columns = [AC_P, AC_E, AC_Y, DC_P]
        results_total = pd.DataFrame(index=weather.index, columns=[AC_P, DC_P]).fillna(0)
        results_total.index.name = 'Time'
        results_kwp = 0
        # results_avg = {}
        energy_price = 0

        for key, configs in self.items():
            if key == 'array':
                result = results[key]
                results_kwp += float(configs.module_parameters['pdc0']) * configs.inverters_per_system * \
                               configs.modules_per_inverter / 1000
                results_total[[AC_P, DC_P]] += result[['p_ac', 'p_dc']].abs().values

                p_pv = results[key]['p_dc'].to_frame()
                p_pv = p_pv.groupby(p_pv.index.hour).mean()
                p_pv = p_pv.rename(columns={'p_dc': 'p_pv'})
                results.update({'daily': p_pv})

        evaluation = False
        key = 'ees'
        if key in self:
            # TODO: Abfrage ob tz-infos vorhanden
            # p_isc.index = p_isc.index.tz_localize('Europe/Berlin')
            # TODO: N.A./null Werte in der Zeitreihe
            evaluation = True
            configs = self[key]
            domestic = pd.read_csv(configs.data_path, delimiter=',', index_col='time', parse_dates=True,
                                   dayfirst=True)
            start = domestic.index[0]
            stop = domestic.index[-1]
            energy_price = configs.energy_price
            domestic.index = (domestic.index + dt.timedelta(minutes=30))[:].floor('H')
            # TODO: delete following line (used because ISC consumption used for domestic consumption)
            # domestic = domestic / 100
            power = domestic.rename(columns={'power': 'p_dom'})

            if results != {}:
                p_pv = results['array'][['p_ac']]
                p_pv.index = p_pv.index.tz_convert(None)
                p_pv.index = p_pv.index[:].floor('H')
                p_pv = p_pv.rename(columns={'p_ac': 'p_pv'})
                power = power.join(p_pv)
            else:
                print('No pv available.')
                power_pv = pd.DataFrame(columns={'p_pv'}, index=power.index, data=np.zeros(power.values.__len__()))
                power = power.join(power_pv)

            p_dom = domestic['power'].to_frame()
            p_dom = p_dom.groupby(p_dom.index.hour).mean()
            p_dom = p_dom.rename(columns={'power': 'p_dom'})
            if results != {}:
                p_dom = pd.concat([p_dom, results['daily']], axis=1, join='inner')
            results.update({'daily': p_dom})

            # energy yield and costs
            pv_yield = power['p_pv'][start:stop].sum()
            consumption = power['p_dom'][start:stop].sum()
            energy_yield = pv_yield - consumption
            energy_balance = (power['p_pv'][start:stop] - power['p_dom'][start:stop]) / 1000

            # energy costs raw
            energy_costs_raw = (-consumption / 1000 * configs.energy_price).sum()

            # energy costs with PV
            energy_costs_pv = 0
            for dW in energy_balance:
                if dW > 0:
                    energy_costs_pv += dW * configs.feed_in_tariff_pv
                elif dW < 0:
                    energy_costs_pv += dW * configs.energy_price

            # energy costs with PV and battery
            energy_costs_bat = 0
            storage = configs.capacity * 0.5
            for dW in energy_balance:
                if dW > 0:
                    if storage < configs.capacity:
                        storage += dW
                    if storage >= configs.capacity:
                        if storage > configs.capacity:
                            dW = storage - configs.capacity
                            storage = configs.capacity
                        energy_costs_bat += dW * configs.feed_in_tariff_pv
                elif dW < 0:
                    if storage > 0:
                        storage += dW
                    if storage <= 0:
                        if storage < 0:
                            dW = storage
                            storage = 0
                        energy_costs_bat += dW * configs.energy_price

            results_ees = {'start': power.index[0],
                           'stop': power.index[-1],
                           'costs_raw': [energy_costs_raw],
                           'costs_pv': [energy_costs_pv],
                           'costs_pv_bat': [energy_costs_bat]}

            results.update({key: pd.DataFrame(data=results_ees)})

        key = 'ev'
        if key in self:
            configs = self[key]
            start = domestic.index[0]
            stop = domestic.index[-1]
            if configs.charge_mode == 'SLOW':
                charge_duration = 10
            elif configs.charge_mode == 'NORMAL':
                charge_duration = 7
            elif configs.charge_mode == 'FAST':
                charge_duration = 4
            else:
                charge_duration = int(configs.charge_mode)
            charge_time = [int(x) for x in configs.charge_time.split('-')]

            data = configs.house_connection_point * 1000 + (results['daily']['p_pv'] - results['daily']['p_dom'])
            data_tmp = data[charge_time[0]:].append(data[:charge_time[1]])

            energy_potential = data_tmp.mean() / 1000 * charge_duration
            energy_potential_min = data_tmp.min() / 1000 * charge_duration
            km_potential = energy_potential / configs.fuel_consumption * 100
            km_potential_min = energy_potential_min / configs.fuel_consumption * 100
            ev_potential = int(km_potential / (configs.driving_distance / 7))
            ev_potential_min = int(km_potential_min / (configs.driving_distance / 7))
            ev_distance_total = int(configs.quantity * configs.driving_distance * (stop - start).days / 7)
            energy_costs_raw = np.round(ev_distance_total / 100 * configs.fuel_consumption * energy_price, 1)

            results_ev = {'energy costs': [energy_costs_raw],
                          'ev potential without LMS': [int(data.min() / 4600)],
                          'energy potential': [energy_potential_min],
                          'km potential': [km_potential_min],
                          'ev potential': [ev_potential_min],
                          'house connection point': [configs.house_connection_point]}
            results.update({key: pd.DataFrame(data=results_ev)})

        results_total[AC_E] = results_total[AC_P] / 1000 * hours
        results_total[AC_Y] = results_total[AC_E] / results_kwp

        results_total[DC_E] = results_total[DC_P] / 1000 * hours

        results_total = results_total.dropna(axis='index', how='all')

        yield_specific = round(results_total[AC_Y].sum(), 2)
        yield_energy = round(results_total[AC_E].sum(), 2)

        # self._database.get(file = ../data/system)

        ghi = round((weather['ghi'] / 1000 * hours).sum(), 2)
        dhi = round((weather['dhi'] / 1000 * hours).sum(), 2)

        results_total = results_total[results_total_columns]

        results_summary_columns = [AC_Y, AC_E]
        results_summary_data = [yield_specific, yield_energy]

        results_summary_columns.append('GHI [kWh/m^2]')
        results_summary_data.append(ghi)

        results_summary_columns.append('DHI [kWh/m^2]')
        results_summary_data.append(dhi)

        results_summary = pd.DataFrame(data=[results_summary_data], columns=results_summary_columns)

        self._write_csv(results_summary, results_total, results)
        for key, configs in self.items():
            if key == 'ees' or key == 'ev':
                self._write_pdf(results_summary, results_total, results)
                os.system(self._results_pdf)
                break
        # self._write_excel(results_summary, results_total, results)
        return {
            'status': 'success',
            'yield_energy': yield_energy,
            'yield_specific': yield_specific
        }

    def _write_csv(self, results_summary, results_total, results):
        for key, configs in self.items():
            configs_name = configs.name
            # for configs_type in self._component_types:
            #     configs_name = configs_name.replace(configs_type, '')
            if len(configs_name) < 1:
                configs_name += str(list(self.values()).index(configs) + 1)

            results[key].to_csv(os.path.join(self._results_dir, 'results_' + configs_name + '.csv'),
                                encoding='utf-8-sig')

        results_total.to_csv(os.path.join(self._results_dir, 'results.csv'), encoding='utf-8-sig')
        results_summary.to_csv(self._results_csv, encoding='utf-8-sig', index=False)

        return results

    def _write_pdf(self, results_summary, results_total, results):
        self._create_graphics(results_summary, results_total, results)
        doc = SimpleDocTemplate(self._results_pdf, pagesize=letter,
                                rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=18)
        try:
            pdf = []
            logo = Image('data/isc_logo.jpg', 60, 60, hAlign='CENTER')
            img_1 = Image(os.path.join(self._results_dir, 'array.png'), 250, 200)
            img_2 = Image(os.path.join(self._results_dir, 'ees_1.png'), 400, 260)
            img_3 = Image(os.path.join(self._results_dir, 'ev_2.png'), 400, 260)
            client_name = "Steffen Friedriszik"
            address_parts = ["Rudolf Diesel Str. 17", "78467 Konstanz"]
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(name='Normal_right', alignment=TA_RIGHT))
            styles.add(ParagraphStyle(name='Normal_center', alignment=TA_CENTER))
            styles.add(ParagraphStyle(name='Heading222', fontName='Helvetica-Bold', fontSize=18, alignment=TA_CENTER,
                                      textColor=colors.blue, leading=20))
            styles.add(ParagraphStyle(name='Heading', fontName='Helvetica-Bold', fontSize=18, alignment=TA_CENTER,
                                      textColor=colors.rgb2cmyk(0, 79, 159), leading=20))
            styles.add(ParagraphStyle(name='Normal_bold', fontSize=14, leading=16))
            title = 'Auswertung der Messergebnisse <br/> für <br/> %s' % self.name
            text_intro = 'Sehr geehrter Herr %s, im Folgenden sehen Sie die Auswertung der Messergebnisse und ihre ' \
                         'Interpretation. Die Analyse erfolgt hinsichtlich allgemeinen, wissenschaftlichen ' \
                         'Aspekten, sowie unter monetären Gesichtspunkten. <br/>' \
                         'Die Messung dauerte %s Tage. Sie startete am %s Uhr und endete am %s Uhr. ' \
                         % (client_name,
                            (results['ees']['stop'][0]-results['ees']['start'][0]).days,
                            results['ees']['start'][0].strftime("%d/%m/%Y, %H:%M:%S"),
                            results['ees']['stop'][0].strftime("%d/%m/%Y, %H:%M:%S"))
            text_img1 = 'In Abbildung 1 sehen Sie Ihre, in dem Messzeitraum erzeugten, Energiekosten im Vergleich ' \
                        'zu den Energiekosten, bei der ' \
                        'angegebenen PV-Anlage und der Speicherkapazität. Die gesamten Energiekosten, ohne PV und ' \
                        'Speicher belaufen sich auf %s €. Wenn Sie eine PV Anlage mit %s Modulen' \
                        'dazu kaufen, sparen sie im Messzeitraum %s €. <br/> ' \
                        'Wenn Sie sich für eine Kombination ' \
                        'aus PV-Anlage und Batteriespeicher mit %s kWh entscheiden, betragen die Energiekosten %s €. ' \
                        'Sie sparen damit %s € im Vergleich zu einem System ohne PV und Speicher. Die berechneten ' \
                        'Preise verstehen sich ohne Investitionskosten.'\
                        % ('%.1f' % -results['ees']['costs_raw'],
                            self['array'].modules_per_inverter,
                            '%.1f' % (-results['ees']['costs_raw'] + results['ees']['costs_pv']),
                            self['ees'].capacity,
                            '%.1f' % -results['ees']['costs_pv_bat'],
                            (results['ees']['costs_pv_bat'][0] - results['ees']['costs_raw'][0]).round(1))
            text_img2 = 'In Abbildung 2 ist ihr Energieverbrauch und die Erzeugung während eines Tages zu erkennen. Ihr ' \
                        'durchschnittlicher Verbrauch liegt bei ca. %s Watt. Ihren maximalen Verbrauch von %s W ' \
                        'haben Sie um %s Uhr. <br/>' \
                        % ('%.1f' % results['daily']['p_dom'].mean(),
                           '%.1f' % results['daily']['p_dom'].max(),
                           results['daily']['p_dom'].idxmax())
            text_img3 = 'Abbildung 3 zeigt die Energiebilanz in Ihrem Haus. Wenn mehr Energie erzeugt wird, wie sie ' \
                     'Verbrauchen, ist die Energiebilanz positiv. Wenn Sie mehr verbrauchen wie erzeugt wird, ist ' \
                     'die Bilanz negativ. Da ihr Hausanschluss für %s kW ausgelegt ist ergibt sich, unter ' \
                     'Berücksichtigung Ihres individuellen Lastprofils und unter Verwendung eines Lademanagement-' \
                     'Systems, ein tägliches Ladepotential von %s kWh. ' \
                     'Das entspricht einer Fahrleistung von %s km. Bei der angegebenen Strecke je Fahrzeug, könnten Sie damit %s ' \
                     'Fahrzeuge betreiben ohne Ihren Hausanschluss zu überlasten. <br/> ' \
                     'Wenn sie kein Lademanagement installieren, könnten sie maximal %s Fahrzeuge im angegebenen ' \
                     'Zeitraum laden, ohne dabei ihren Hausanschluss zu überlasten. (Vorraussetzung der Gleichzeitigkeit beim Ladevorgang)' \
                     % (results['ev']['house connection point'][0],
                        '%.1f' % results['ev']['energy potential'][0],
                        '%.1f' % results['ev']['km potential'][0],
                        results['ev']['ev potential'][0],
                        results['ev']['ev potential without LMS'][0])
            text_remarks = 'Die oben aufgeführten Berechnungen werden nach besten Gewissen aufgrund der eingegebenen ' \
                           'Angaben berechnet. Größen wie die Ladezeit variieren in der realität täglich, werden ' \
                           'hier jedoch als konstant angenommen.'

            pdf.append(logo)
            pdf.append(Paragraph('%s' % dt.date.today(), styles["Normal_right"]))
            pdf.append(Spacer(1, 12))
            pdf.append(Paragraph(client_name, styles["Normal"]))
            for part in address_parts:
                ptext = '%s' % part.strip()
                pdf.append(Paragraph(ptext, styles["Normal"]))
            pdf.append(Spacer(1, 12))
            pdf.append(Paragraph(title, styles["Heading"]))
            pdf.append(Spacer(1, 24))
            pdf.append(Paragraph(text_intro, styles["Normal"]))
            pdf.append(Spacer(1, 12))
            pdf.append(Paragraph(text_img1, styles["Normal"]))
            pdf.append(Spacer(1, 12))
            pdf.append(img_1)
            pdf.append(Spacer(1, 12))
            pdf.append(Paragraph(text_img2, styles["Normal"]))
            pdf.append(Spacer(1, 12))
            # pdf.append(Table([[img_2, img_3]], colWidths=[275, 275], rowHeights=[220]))
            pdf.append(img_2)
            pdf.append(Spacer(1, 12))
            pdf.append(Paragraph(text_img3, styles["Normal"]))
            pdf.append(Spacer(1, 12))
            pdf.append(img_3)
            pdf.append(Spacer(1, 12))
            pdf.append(Paragraph('Bemerkungen:', styles["Normal_bold"]))
            pdf.append(Paragraph(text_remarks, styles["Normal"]))
            pdf.append(Spacer(1, 24))
            pdf.append(
                Paragraph('mit freundlichen Grüßen und eine sonnige Zukunft wünscht Ihnen ihr ', styles["Normal"]))
            pdf.append(Spacer(1, 12))
            pdf.append(Paragraph('ISC Konstanz', styles["Normal_bold"]))
            pdf.append(Spacer(1, 12))
            doc.build(pdf)
        except ValueError:
            print('PDF not created %s' % ValueError)
        print('PDF created successfully')

    def _create_graphics(self, results_summary, results_total, results):
        size_large = 30
        size_medium = 27
        size_small = 24
        width = 0.6
        for key, configs in self.items():
            if key == 'array':
                fig1, ax1 = plt.subplots(figsize=(10, 10))
                data = [
                    -results['ees']['costs_raw'].values[0].round(1),
                    -results['ees']['costs_pv'].values[0].round(1),
                    -results['ees']['costs_pv_bat'].values[0].round(1),
                ]
                plot1 = ax1.bar(np.arange(3), data, width)
                # ax1.set_title('Energy costs in comparison', fontsize=size_large)
                ax1.set_ylabel('energy costs in €', fontsize=size_medium)
                ax1.tick_params(axis='y', labelsize=size_medium)
                ax1.set_xticks(np.arange(3))
                label = ['RAW', 'mit PV', 'mit PV und EES']
                ax1.set_xticklabels(label, fontsize=size_medium)
                ax1.yaxis.set_ticks_position('left')
                ax1.xaxis.set_ticks_position('bottom')
                ax1.spines['right'].set_visible(False)
                ax1.spines['top'].set_visible(False)
                ax1.bar_label(plot1, padding=3, fontsize=size_small)
                fig1.tight_layout()
                plt.savefig(self._results_dir + '\\' + key)

            if key == 'ees':
                fig2, ax2 = plt.subplots(figsize=(23, 15))
                # ax2.set_title('energy', fontsize=size_large)
                ax2.plot(results['daily'], linewidth=2.5)
                ax2.set_xlim([0, 23])
                ax2.set_xlabel('hour of the day', fontsize=size_medium)
                ax2.set_ylabel('power [KW]', fontsize=size_medium)
                ax2.legend(results['daily'].columns, fontsize=size_small)
                ax2.grid(which='both')
                ax2.grid(b=True, which='major', color='grey', linestyle='-')
                ax2.grid(b=True, which='minor', color='grey', linestyle='--', alpha=0.2)
                ax2.minorticks_on()
                ax2.tick_params(axis='y', labelsize=size_small)
                ax2.tick_params(axis='x', labelsize=size_small)
                plt.savefig(self._results_dir + '\\' + key + '_1')

            if key == 'ev':
                fig3, ax3 = plt.subplots(figsize=(23, 15))
                ax3.plot(results['daily']['p_pv'] - results['daily']['p_dom'], linewidth=2.5)
                ax3.set_xlim([0, 23])
                ax3.set_xlabel('hour of the day', fontsize=size_medium)
                ax3.set_ylabel('power [KW]', fontsize=size_medium)
                ax3.legend(['power balance'], fontsize=size_small)
                ax3.grid(which='both')
                ax3.grid(b=True, which='major', color='grey', linestyle='-')
                ax3.grid(b=True, which='minor', color='grey', linestyle='--', alpha=0.2)
                ax3.minorticks_on()
                ax3.tick_params(axis='y', labelsize=size_small)
                ax3.tick_params(axis='x', labelsize=size_small)
                plt.savefig(self._results_dir + '\\' + key + '_2')

    def _write_excel(self, results_summary, results_total, results):
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Border, Side

            border_side = Side(border_style=None)
            border = Border(top=border_side,
                            right=border_side,
                            bottom=border_side,
                            left=border_side)

            results_book = Workbook()
            results_writer = pd.ExcelWriter(self._results_excel, engine='openpyxl', engine_kwargs={'encoding': 'utf-8-sig'})
            results_writer.book = results_book
            results_summary.to_excel(results_writer, sheet_name='Summary', float_format="%.2f", encoding='utf-8-sig')
            results_book['Summary'].delete_cols(1, 1)
            results_book.remove_sheet(results_book.active)
            results_book.active = 0

            results_total.tz_localize(None).to_excel(results_writer, sheet_name=self.name, encoding='utf-8-sig')

            for key, configs in self.items():
                configs_name = configs.name
                for configs_type in self._component_types:
                    configs_name = configs_name.replace(configs_type, '')
                if len(configs_name) < 1:
                    configs_name += str(list(self.values()).index(configs) + 1)

                results[key].tz_localize(None).to_excel(results_writer,
                                                        sheet_name=self.name + ' ' + configs_name,
                                                        encoding='utf-8-sig')

            results_summary.to_excel(results_writer, sheet_name='Summary', float_format="%.2f", encoding='utf-8-sig')
            results_total.tz_localize(None).to_excel(results_writer, sheet_name=self.name, encoding='utf-8-sig')
            results_book['Summary'].delete_cols(1, 1)
            results_sheets = results_book._sheets
            results_sheets.insert(0, results_sheets.pop(len(results_sheets) - 1))
            results_sheets.insert(0, results_sheets.pop(len(results_sheets) - 1))

            for result_sheet in results_book:
                result_index_width = 0
                for result_row in result_sheet:
                    result_row[0].border = border
                    result_index_width = max(result_index_width, len(str(result_row[0].value)))

                result_sheet.column_dimensions[get_column_letter(1)].width = result_index_width + 2

                for result_column in range(1, len(result_sheet[1])):
                    result_column_width = len(str(result_sheet[1][result_column].value))
                    result_sheet[1][result_column].border = border
                    result_sheet.column_dimensions[get_column_letter(result_column+1)].width = result_column_width + 2

            results_book.save(self._results_excel)
            results_writer.close()

        except ImportError as e:
            logger.debug("Unable to write excel file for {} of system {}: {}".format(
                os.path.abspath(self._results_excel), self.name, str(e)))
